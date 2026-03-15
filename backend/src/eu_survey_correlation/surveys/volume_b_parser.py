"""Extract answer distributions from Eurobarometer Volume B Excel files.

Reads questions from michlou_survey_tri.csv, finds the corresponding sheet in each
Volume B Excel file, and extracts answer counts/percentages broken down by:
- EU27 total
- Income difficulty (poor / medium / rich)
- Class belonging (working_class / lower_middle / middle / upper_middle / upper)
- Gender (male / female)
- Occupation (self_employed / manager / employee / manual_worker / etc.)

Output: data/surveys/volume_b_answer_distributions.csv (long format)
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
from loguru import logger
from tqdm import tqdm

DATA_PATH = Path("./data")
VOLUME_B_DIR = DATA_PATH / "volume_b_docs"
SURVEY_CSV = DATA_PATH / "surveys" / "michlou_survey_tri.csv"
OUTPUT_CSV = DATA_PATH / "surveys" / "volume_b_answer_distributions.csv"

# ---------------------------------------------------------------------------
# Keyword → normalized demographic value mappings
# ---------------------------------------------------------------------------

INCOME_DIFFICULTY_MAP = {
    "most of the time": "poor",
    "plupart du temps": "poor",
    "time to time": "medium",
    "temps en temps": "medium",
    "almost never": "rich",
    "pratiquement jamais": "rich",
    "presque jamais": "rich",
}

CLASS_BELONGING_MAP = {
    "ouvrière": "working_class",
    "working class": "working_class",
    "inférieure": "lower_middle",
    "lower middle": "lower_middle",
    "supérieure": "upper_middle",
    "upper middle": "upper_middle",
    "plus élevée": "upper",
    "upper class": "upper",
    "highest": "upper",
}

GENDER_MAP = {
    "femme": "female",
    "female": "female",
    "woman": "female",
    "male": "male",
    "homme": "male",
    "man": "male",
}


OCCUPATION_MAP = {
    "indépendant": "self_employed",
    "self-employed": "self_employed",
    "self- employed": "self_employed",
    "cadre": "manager",
    "manager": "manager",
    "autre": "other_white_collar",
    "other white": "other_white_collar",
    "employee": "employee",
    "ouvrier": "manual_worker",
    "manual worker": "manual_worker",
    "foyer": "house_person",
    "house person": "house_person",
    "chômeur": "unemployed",
    "unemployed": "unemployed",
    "retraité": "retired",
    "retired": "retired",
    "étudiant": "student",
    "student": "student",
    "not working": "not_working",
}


def _normalize(text: str | None) -> str:
    """Lowercase, strip, collapse whitespace."""
    if not text:
        return ""
    return " ".join(str(text).lower().strip().split())


def _match_keyword(text: str, mapping: dict[str, str]) -> str | None:
    """Return the first matching value from a keyword→label mapping."""
    t = _normalize(text)
    for keyword, label in mapping.items():
        if keyword in t:
            return label
    return None


class VolumeBParser:
    """Parses Eurobarometer Volume B cross-tabulation Excel files."""

    def __init__(self, volume_b_dir: Path = VOLUME_B_DIR):
        self.volume_b_dir = volume_b_dir

    # ------------------------------------------------------------------
    # Format detection
    # ------------------------------------------------------------------

    def detect_format(self, rows: list[tuple]) -> dict:
        """Auto-detect header layout by searching for 'Gender'/'Sexe' keyword."""
        cat_header_row = -1

        for ri in range(min(15, len(rows))):
            for ci, v in enumerate(rows[ri]):
                t = _normalize(v)
                if not t:
                    continue
                # Category headers are short (< 80 chars) — skip long question text
                if len(t) > 80:
                    continue
                if ci >= 2 and ("gender" in t or "sexe" in t):
                    cat_header_row = ri
                    break
            if cat_header_row >= 0:
                break

        if cat_header_row < 0:
            # Fallback: look for "EU27" or "UE27" in sub-header rows
            for ri in range(min(15, len(rows))):
                for ci, v in enumerate(rows[ri]):
                    t = _normalize(v)
                    if not t or len(t) > 80:
                        continue
                    if "eu27" in t or "ue27" in t:
                        cat_header_row = ri - 1
                        break
                if cat_header_row >= 0:
                    break

        if cat_header_row < 0:
            return {"format": "unknown", "cat_header_row": -1}

        sub_header_row = cat_header_row + 1

        # Detect standard vs flash
        # Standard: bilingual text with \n-\n, typically row 7-8
        # Flash: English-only, typically row 10-11, has stat-test letter row at +2
        is_flash = cat_header_row >= 9
        if not is_flash:
            # Double-check: standard files have bilingual "\n-\n" in header cells
            has_bilingual = False
            for v in rows[cat_header_row]:
                if v and "\n" in str(v):
                    has_bilingual = True
                    break
            if not has_bilingual:
                # Also check sub-header row
                for v in rows[sub_header_row]:
                    if v and "\n" in str(v):
                        has_bilingual = True
                        break
            if not has_bilingual:
                is_flash = True

        fmt = "flash" if is_flash else "standard"

        # Find base row(s) — prefer the "weighted" base over plain "total"
        base_row = -1
        fallback_base_row = -1
        for ri in range(sub_header_row + 1, min(sub_header_row + 10, len(rows))):
            first_cell = _normalize(rows[ri][0]) if rows[ri][0] else ""
            second_cell = (
                _normalize(rows[ri][1]) if len(rows[ri]) > 1 and rows[ri][1] else ""
            )
            combined = first_cell + " " + second_cell
            if "weighted" in combined:
                base_row = ri
                break
            if fallback_base_row < 0 and (
                "base:" in combined or combined.strip() == "total"
            ):
                fallback_base_row = ri

        if base_row < 0:
            base_row = (
                fallback_base_row if fallback_base_row >= 0 else sub_header_row + 1
            )

        # Data rows start after the last base-like row
        # Scan forward to find the first actual answer row
        label_col = 1 if fmt == "standard" else 0
        data_start_row = base_row + 1
        for ri in range(data_start_row, min(data_start_row + 8, len(rows))):
            row = rows[ri]
            if label_col >= len(row):
                continue
            cell = _normalize(row[label_col])
            if not cell:
                continue
            # Skip any remaining base/header rows
            if "base" in cell or "weighted" in cell:
                continue
            data_start_row = ri
            break

        return {
            "format": fmt,
            "cat_header_row": cat_header_row,
            "sub_header_row": sub_header_row,
            "base_row": base_row,
            "data_start_row": data_start_row,
        }

    # ------------------------------------------------------------------
    # Column discovery
    # ------------------------------------------------------------------

    def discover_columns(self, cat_row: tuple, sub_row: tuple, base_row: tuple) -> dict:
        """Find demographic columns by keyword matching in header rows."""
        total_col = -1
        demographics: list[dict] = []

        # Build category spans: for each col, determine which category group it belongs to
        # Categories in cat_row are sparse (only first col of each group has a value)
        current_cat = ""
        cat_spans: dict[int, str] = {}
        for ci in range(len(cat_row)):
            v = _normalize(cat_row[ci])
            if v:
                current_cat = v
            cat_spans[ci] = current_cat

        for ci in range(len(sub_row)):
            sub_val = _normalize(sub_row[ci])
            if not sub_val:
                continue

            cat_val = cat_spans.get(ci, "")

            # EU27 total column
            if "eu27" in sub_val or "ue27" in sub_val:
                total_col = ci
                continue

            # Income difficulty
            if "difficul" in cat_val:
                label = _match_keyword(sub_val, INCOME_DIFFICULTY_MAP)
                if label:
                    demographics.append(
                        {"type": "income_difficulty", "value": label, "col_idx": ci}
                    )
                continue

            # Class belonging
            if "appartenir" in cat_val or "belonging" in cat_val:
                # Special handling: "middle class" without "lower" or "upper" qualifiers
                label = _match_keyword(sub_val, CLASS_BELONGING_MAP)
                if label is None and (
                    "moyenne" in sub_val or "middle class" in sub_val
                ):
                    label = "middle"
                if label:
                    demographics.append(
                        {"type": "class_belonging", "value": label, "col_idx": ci}
                    )
                continue

            # Gender
            if "sexe" in cat_val or "gender" in cat_val:
                label = _match_keyword(sub_val, GENDER_MAP)
                if label:
                    demographics.append(
                        {"type": "gender", "value": label, "col_idx": ci}
                    )
                continue

            # Occupation
            if "socio" in cat_val or "occupation" in cat_val:
                label = _match_keyword(sub_val, OCCUPATION_MAP)
                if label:
                    demographics.append(
                        {"type": "occupation", "value": label, "col_idx": ci}
                    )
                continue

        return {
            "total_col": total_col,
            "demographics": demographics,
            "base_row_data": base_row,
        }

    # ------------------------------------------------------------------
    # Answer row parsing
    # ------------------------------------------------------------------

    def parse_answer_rows(self, rows: list[tuple], format_info: dict) -> list[dict]:
        """Identify answer rows and pair them with their percentage rows."""
        fmt = format_info["format"]
        start = format_info["data_start_row"]
        label_col = 1 if fmt == "standard" else 0
        answers = []

        ri = start
        while ri < len(rows):
            row = rows[ri]
            if label_col >= len(row):
                ri += 1
                continue

            label = str(row[label_col]).strip() if row[label_col] else ""
            if not label:
                ri += 1
                continue

            # Clean bilingual labels: "FR text\nEN text" → keep EN
            if "\n" in label:
                parts = label.split("\n")
                # Take the last non-empty part (usually EN)
                label = parts[-1].strip()

            # Detect summary rows — "Total 'Agree'" etc., not "Totally agree"
            label_lower = label.lower()
            is_summary = label_lower.startswith("total ") or label_lower == "total"

            # The next row should contain percentages
            pct_row_idx = ri + 1
            if pct_row_idx >= len(rows):
                break

            # For flash format, skip stat-test letter rows
            step = 2 if fmt == "standard" else 3

            answers.append(
                {
                    "answer_label": label,
                    "is_summary": is_summary,
                    "count_row_idx": ri,
                    "pct_row_idx": pct_row_idx,
                }
            )

            ri += step

        return answers

    # ------------------------------------------------------------------
    # Sheet extraction
    # ------------------------------------------------------------------

    def _safe_float(self, val) -> float | None:
        """Convert a cell value to float, returning None on failure."""
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _normalize_pct(self, pct_val: float | None) -> float | None:
        """Normalize percentage to 0-100 scale."""
        if pct_val is None:
            return None
        # Values <= 1.0 are ratios (0.15 = 15%), values > 1 are already percentages
        if 0 < pct_val <= 1.0:
            return round(pct_val * 100, 2)
        return round(pct_val, 2)

    def extract_sheet(
        self,
        rows: list[tuple],
        sheet_name: str,
        file_name: str,
        question_clean: str,
    ) -> list[dict]:
        """Extract all answer×demographic data from a single sheet."""
        format_info = self.detect_format(rows)

        if format_info["cat_header_row"] < 0:
            logger.warning(f"Could not detect format for {file_name}/{sheet_name}")
            return []

        cat_row = rows[format_info["cat_header_row"]]
        sub_row = rows[format_info["sub_header_row"]]
        base_row_data = rows[format_info["base_row"]]

        columns = self.discover_columns(cat_row, sub_row, base_row_data)
        answers = self.parse_answer_rows(rows, format_info)

        if not answers:
            logger.warning(f"No answers found in {file_name}/{sheet_name}")
            return []

        results = []

        # Build list of all demographic slices to extract
        slices = []
        if columns["total_col"] >= 0:
            slices.append(
                {
                    "type": "total",
                    "value": "eu27",
                    "col_idx": columns["total_col"],
                }
            )
        slices.extend(columns["demographics"])

        for answer in answers:
            count_row = rows[answer["count_row_idx"]]
            pct_row = rows[answer["pct_row_idx"]]

            for slc in slices:
                ci = slc["col_idx"]
                count_val = self._safe_float(
                    count_row[ci] if ci < len(count_row) else None
                )
                pct_val = self._safe_float(pct_row[ci] if ci < len(pct_row) else None)
                base_val = self._safe_float(
                    base_row_data[ci] if ci < len(base_row_data) else None
                )

                pct_normalized = self._normalize_pct(pct_val)

                results.append(
                    {
                        "sheet_id": sheet_name,
                        "file_name": file_name,
                        "question_clean": question_clean,
                        "answer_label": answer["answer_label"],
                        "is_summary": answer["is_summary"],
                        "demographic_type": slc["type"],
                        "demographic_value": slc["value"],
                        "count": count_val,
                        "pct": pct_normalized,
                        "total_base": base_val,
                    }
                )

        return results

    # ------------------------------------------------------------------
    # Main pipeline
    # ------------------------------------------------------------------

    def extract_all(self, survey_csv: Path = SURVEY_CSV) -> pd.DataFrame:
        """Extract answer distributions for all questions in the survey CSV."""
        df_surveys = pd.read_csv(survey_csv)
        logger.info(f"Loaded {len(df_surveys)} questions from {survey_csv.name}")

        # Group by file to open each workbook only once
        grouped = df_surveys.groupby("file_name")
        all_results: list[dict] = []

        for file_name, group in tqdm(grouped, desc="Processing files"):
            fpath = self.volume_b_dir / file_name
            if not fpath.exists():
                logger.warning(f"File not found: {file_name}")
                continue

            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            except Exception as e:
                logger.error(f"Cannot open {file_name}: {e}")
                continue

            for _, row in group.iterrows():
                sheet_id = str(row["sheet_id"]).strip()
                question_clean = str(row.get("question_clean", ""))

                # Try exact match, then zero-padded (e.g. "10" → "0010")
                actual_sheet = None
                if sheet_id in wb.sheetnames:
                    actual_sheet = sheet_id
                else:
                    # Try zero-padded variants
                    for sn in wb.sheetnames:
                        if sn.lstrip("0") == sheet_id or sn == sheet_id.zfill(4):
                            actual_sheet = sn
                            break
                if actual_sheet is None:
                    logger.warning(f"Sheet '{sheet_id}' not found in {file_name}")
                    continue

                ws = wb[actual_sheet]
                sheet_rows = list(ws.iter_rows(values_only=True))

                results = self.extract_sheet(
                    sheet_rows, sheet_id, file_name, question_clean
                )
                all_results.extend(results)

            wb.close()

        df_out = pd.DataFrame(all_results)
        logger.info(
            f"Extracted {len(df_out)} rows "
            f"({df_out['sheet_id'].nunique()} questions, "
            f"{df_out['file_name'].nunique()} files)"
        )
        return df_out
